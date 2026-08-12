# -*- coding: utf-8 -*-
"""
ETL NETTOYAGE — v2.4
=====================
Reads raw yearly Excel bilans, cleans & standardises them,
imputes missing values (local then global), and exports one
clean_{year}.xlsx per year with two sheets:
  • Formations        — one row per formation
  • Participants_Detail — one row per participant

Usage:
    pip install pandas openpyxl
    python etl_clean.py
"""

import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────

RAW_FOLDER   = Path(r"C:\Users\linab\OneDrive\Documents\data")
CLEAN_FOLDER = Path(r"C:\Users\linab\OneDrive\Documents\cleaned_bilans")
CLEAN_FOLDER.mkdir(exist_ok=True)

MODES_VALIDES = ["FPP", "FIP", "FIAE", "FD", "FIAI", "FIS", "FIT", "FII", "FE"]

STANDARD_COLUMNS = [
    "formation_id", "year", "mode", "theme", "montant", "montant_raw",
    "date_debut", "date_fin", "nb_jours_calc", "nb_heures",
    "bureau_formation", "formateur", "facture",
    "etat_dossier", "status", "source_file",
]

SHEET_KEYWORDS = ["inventaire", "bilan", "synthese", "synthèse", "recap", "récap"]

EXPECTED_COLS = ["mode", "num", "thème", "theme", "formateur", "montant",
                 "date", "etat", "facture", "bureau"]

AFFECTATIONS_CONNUES = {
    "REGROW"            : "RÉSEAU DES AGENCES",
    "RESEAU"            : "RÉSEAU DES AGENCES",
    "RESEAU DES AGENCES": "RÉSEAU DES AGENCES",
    "CHEFS D'AGENCES"   : "CHEFS D'AGENCES",
    "SIEGE"             : "SIÈGE",
    "INTERNATIONAL"     : "INTERNATIONAL",
    "COMMERCIALE"       : "COMMERCIALE",
    "DIRECTION"         : "DIRECTION GÉNÉRALE",
}

ETAT_MAP = {
    "complet"          : "Complet",
    "incomplet"        : "Incomplet",
    "clôturé"          : "Clôturé",
    "cloture"          : "Clôturé",
    "non ristournable" : "Non Ristournable",
    "en cours"         : "En cours",
    "non renseigné"    : "Non renseigné",
}


# ─────────────────────────────────────────────
# HELPERS — DATA CLEANING
# ─────────────────────────────────────────────

def clean_date(val) -> pd.Timestamp:
    """
    Convertit une valeur brute en Timestamp pandas.

    Tente d'interpréter `val` comme une date via pd.to_datetime.
    Les dates "sentinelles" aberrantes issues d'Excel (1899-01-xx,
    1900-01-xx, 1970-01-xx) sont considérées comme invalides et
    retournées sous forme de NaT (Not a Time).

    Paramètres
    ----------
    val : any
        Valeur brute lue depuis Excel (string, datetime, float, etc.)

    Retourne
    --------
    pd.Timestamp | pd.NaT
        La date parsée, ou NaT si la valeur est manquante/aberrante.
    """
    dt = pd.to_datetime(val, errors="coerce")
    if pd.isna(dt):
        return pd.NaT
    if dt.year in (1899, 1900, 1970) and dt.month == 1:
        return pd.NaT
    return dt


def clean_montant(val) -> tuple:
    """
    Nettoie et extrait un montant financier depuis une valeur brute.

    Gère les cas particuliers : valeurs manquantes, libellés textuels
    ("gratuit", "annulé", "non remboursable") et chaînes numériques
    contenant des séparateurs de milliers ou virgules décimales.

    Paramètres
    ----------
    val : any
        Valeur brute du montant lue depuis Excel.

    Retourne
    --------
    tuple (float | None, str | None)
        - Premier élément  : montant numérique nettoyé, ou None si non parsable.
        - Deuxième élément : chaîne brute originale (pour traçabilité), ou None.
    """
    raw = str(val).strip() if pd.notna(val) else ""
    if not raw or raw.lower() == "nan":
        return None, None
    lower = raw.lower()
    if any(w in lower for w in ["gratuit", "annule", "annulé", "non remboursable"]):
        return None, raw
    try:
        cleaned = re.sub(r"[^0-9.,]", "", raw).replace(",", ".")
        return float(cleaned), raw
    except (ValueError, TypeError):
        return None, raw


def compute_status(date_debut, date_fin) -> str | None:
    """
    Détermine le statut d'une formation en fonction de ses dates et de la date du jour.

    Règles appliquées :
      - Si l'une des deux dates est manquante → None (sera imputé plus tard).
      - Si aujourd'hui est compris entre date_debut et date_fin → "En cours".
      - Si date_fin est dépassée → "Terminée".
      - Sinon → "Planifiée".

    Paramètres
    ----------
    date_debut : pd.Timestamp | pd.NaT
        Date de début de la formation.
    date_fin : pd.Timestamp | pd.NaT
        Date de fin de la formation.

    Retourne
    --------
    str | None
        Le statut calculé, ou None si les dates sont insuffisantes.
    """
    today = pd.Timestamp.today().normalize()
    if pd.isna(date_debut) or pd.isna(date_fin):
        return None
    if date_debut <= today <= date_fin:
        return "En cours"
    if date_fin < today:
        return "Terminée"
    return "Planifiée"


def normalize_affectation(val) -> str | None:
    """
    Normalise une valeur d'affectation ou de bureau de formation.

    Convertit la valeur en majuscules puis la recherche dans le
    dictionnaire AFFECTATIONS_CONNUES. Si elle y figure, retourne
    la forme canonique ; sinon, retourne la valeur telle quelle
    (après strip). Les valeurs vides ou NaN sont retournées sans
    modification.

    Paramètres
    ----------
    val : any
        Valeur brute de l'affectation.

    Retourne
    --------
    str | None
        La valeur normalisée ou la valeur originale si non reconnue.
    """
    if pd.isna(val) or not str(val).strip():
        return val
    s = str(val).strip()
    return AFFECTATIONS_CONNUES.get(s.upper(), s)


def normalize_etat_dossier(val) -> str:
    """
    Normalise la valeur de l'état du dossier de formation.

    Règles :
      - Valeur None ou NaN → "Non renseigné".
      - Chaîne vide ou "nan" → "Non renseigné".
      - Valeur purement numérique (ex : "2023" issu d'une erreur de saisie) → "Non renseigné".
      - Sinon, applique le mapping ETAT_MAP (insensible à la casse) ; si absent du mapping,
        retourne la valeur telle quelle.

    Paramètres
    ----------
    val : any
        Valeur brute de l'état du dossier.

    Retourne
    --------
    str
        La valeur normalisée, toujours une chaîne non vide.
    """
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return "Non renseigné"
    s = str(val).strip()
    if not s or s.lower() == "nan":
        return "Non renseigné"
    if re.fullmatch(r"\d{1,4}", s):
        return "Non renseigné"
    return ETAT_MAP.get(s.lower(), s)


def validate_mode(mode, formation_id: str = "") -> str:
    """
    Valide que le mode de formation fait partie de la liste MODES_VALIDES.

    Si le mode est inconnu, affiche un avertissement dans la console et
    retourne la valeur par défaut "FIP".

    Paramètres
    ----------
    mode : any
        Valeur brute du mode de formation.
    formation_id : str, optionnel
        Identifiant de la formation (utilisé dans le message d'avertissement).

    Retourne
    --------
    str
        Le mode validé ou "FIP" si inconnu.
    """
    mode = str(mode).strip()
    if mode not in MODES_VALIDES:
        print(f"   Mode inconnu: '{mode}' (formation {formation_id}) → remplacé par 'FIP'")
        return "FIP"
    return mode


def mode_most_frequent(series: pd.Series):
    """
    Retourne la valeur la plus fréquente (mode statistique) d'une Series pandas.

    Les valeurs nulles sont ignorées. Si la Series est entièrement vide
    ou nulle, retourne None.

    Paramètres
    ----------
    series : pd.Series
        Série de valeurs à analyser.

    Retourne
    --------
    any | None
        La valeur la plus fréquente, ou None si la série est vide.
    """
    m = series.dropna().mode()
    return m.iloc[0] if not m.empty else None


def get_col(row: dict, *candidates):
    """
    Recherche insensible à la casse d'une colonne dans un dictionnaire de ligne.

    Parcourt les noms de colonnes candidats dans l'ordre fourni et retourne
    la valeur du premier nom trouvé dans le dictionnaire (comparaison en
    minuscules après strip). Utile pour gérer les variantes orthographiques
    de noms de colonnes entre différents fichiers Excel.

    Paramètres
    ----------
    row : dict
        Dictionnaire représentant une ligne du DataFrame (clé = nom de colonne).
    *candidates : str
        Noms de colonnes à rechercher, dans l'ordre de priorité.

    Retourne
    --------
    any | None
        La valeur trouvée pour le premier nom candidat correspondant, ou None.
    """
    row_lower = {k.lower().strip(): v for k, v in row.items()}
    for c in candidates:
        val = row_lower.get(c.lower().strip())
        if val is not None:
            return val
    return None


# ─────────────────────────────────────────────
# HELPERS — PARTICIPANT PARSING
# ─────────────────────────────────────────────

def parse_participants(cell_participants, cell_affectation=None) -> list[dict]:
    """
    Parse une cellule multi-lignes de participants en une liste de dictionnaires.

    Chaque participant est identifié par un motif "matricule nom" (ex: "12345 DUPONT Jean").
    La cellule d'affectation est traitée selon ces règles de propagation :
      - Une seule affectation présente → appliquée à tous les participants.
      - Plusieurs affectations → alignées par index avec les participants.
      - Affectation intégrée dans le champ nom → extraite via AFFECTATIONS_CONNUES.

    Paramètres
    ----------
    cell_participants : any
        Contenu brut de la cellule participants (multi-lignes possible).
    cell_affectation : any, optionnel
        Contenu brut de la cellule affectation correspondante.

    Retourne
    --------
    list[dict]
        Liste de dictionnaires avec les clés : 'matricule', 'nom', 'affectation'.
        Retourne une liste vide si la cellule est vide ou ne contient pas de pattern valide.
    """
    if pd.isna(cell_participants):
        return []

    pattern = r"(\d{3,5})\s+(.+)"
    matches  = re.findall(pattern, str(cell_participants))
    if not matches:
        return []

    # Parse affectation cell into clean lines
    affectations: list[str] = []
    if cell_affectation is not None and not pd.isna(cell_affectation):
        for line in str(cell_affectation).replace("\r", "").split("\n"):
            line = line.strip()
            if not line:
                continue
            line_clean = re.sub(r"^\d{3,5}\s+", "", line).strip()
            if line_clean:
                affectations.append(line_clean)

    single_affectation = affectations[0] if len(affectations) == 1 else None

    def split_nom_affectation(nom_raw: str) -> tuple[str, str | None]:
        """
        Sépare le nom du participant de l'affectation éventuellement collée à la fin.

        Vérifie si le nom brut se termine par une clé connue de AFFECTATIONS_CONNUES
        (comparaison en majuscules). Si c'est le cas, extrait et retourne séparément
        le nom nettoyé et l'affectation canonique correspondante.

        Paramètres
        ----------
        nom_raw : str
            Chaîne brute pouvant contenir "NOM PRENOM AFFECTATION".

        Retourne
        --------
        tuple (str, str | None)
            - Nom nettoyé.
            - Affectation extraite (forme canonique) ou None si non trouvée.
        """
        nom_raw = nom_raw.strip()
        for aff_key in AFFECTATIONS_CONNUES:
            if nom_raw.upper().endswith(aff_key):
                nom_clean = nom_raw[: -len(aff_key)].strip()
                return nom_clean, AFFECTATIONS_CONNUES[aff_key]
        return nom_raw, None

    result = []
    for i, (matricule, nom_raw) in enumerate(matches):
        nom_clean, aff_from_nom = split_nom_affectation(nom_raw)

        if single_affectation is not None:
            aff = normalize_affectation(single_affectation)
        else:
            aff_raw = affectations[i] if i < len(affectations) else aff_from_nom
            aff     = normalize_affectation(aff_raw)

        result.append({
            "matricule"  : matricule,
            "nom"        : nom_clean,
            "affectation": aff,
        })

    return result


# ─────────────────────────────────────────────
# SHEET / HEADER DETECTION
# ─────────────────────────────────────────────

def _sheet_col_score(sheet_cols: list) -> int:
    """
    Calcule un score de pertinence pour une liste de colonnes d'une feuille Excel.

    Compare les colonnes (en minuscules) aux noms attendus de EXPECTED_COLS
    et compte le nombre de correspondances (correspondance partielle via `in`).
    Utilisée pour identifier automatiquement la feuille de données principale
    quand plusieurs feuilles sont présentes.

    Paramètres
    ----------
    sheet_cols : list
        Liste des noms de colonnes ou valeurs d'en-tête d'une feuille.

    Retourne
    --------
    int
        Nombre de colonnes attendues trouvées dans la liste.
    """
    cols_lower = [str(c).lower().strip() for c in sheet_cols]
    return sum(1 for exp in EXPECTED_COLS if any(exp in c for c in cols_lower))


def detect_best_sheet(filepath: Path) -> str:
    """
    Identifie automatiquement la feuille de données principale d'un classeur Excel.

    Stratégie en trois étapes :
      1. Si une seule feuille existe, elle est retournée directement.
      2. Si un nom de feuille contient un mot-clé de SHEET_KEYWORDS (ex: "bilan",
         "synthese"), elle est retournée immédiatement.
      3. Sinon, les 5 premières lignes de chaque feuille sont analysées via
         _sheet_col_score ; la feuille obtenant le meilleur score est retournée.

    Paramètres
    ----------
    filepath : Path
        Chemin vers le fichier Excel (.xlsx) à analyser.

    Retourne
    --------
    str
        Nom de la feuille sélectionnée.
    """
    wb = load_workbook(filepath, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    if len(sheet_names) == 1:
        return sheet_names[0]

    for name in sheet_names:
        if any(kw in name.lower() for kw in SHEET_KEYWORDS):
            print(f"    Feuille choisie par nom : '{name}'")
            return name

    best_name, best_score = sheet_names[0], 0
    for name in sheet_names:
        try:
            df_probe = pd.read_excel(filepath, sheet_name=name, header=None, nrows=5)
            for _, row in df_probe.iterrows():
                score = _sheet_col_score(list(row.values))
                if score > best_score:
                    best_score, best_name = score, name
        except Exception:
            continue

    print(f"    Feuille choisie par score ({best_score} cols matchées) : '{best_name}'")
    return best_name


def detect_header_row(filepath: Path, sheet_name: str) -> int:
    """
    Détecte automatiquement la ligne d'en-tête dans une feuille Excel.

    Lit les 15 premières lignes et attribue un score à chacune en comptant
    le nombre de mots-clés de colonnes attendus ("MODE", "THEME", "DATE",
    "MONTANT", etc.) présents dans la ligne. La ligne avec le score le plus
    élevé est considérée comme l'en-tête.

    Paramètres
    ----------
    filepath : Path
        Chemin vers le fichier Excel.
    sheet_name : str
        Nom de la feuille à analyser.

    Retourne
    --------
    int
        Index 0-based de la ligne d'en-tête détectée.
    """
    df_probe = pd.read_excel(filepath, sheet_name=sheet_name, header=None, nrows=15)
    keywords = ["MODE", "NUM", "THÈME", "THEME", "DATE", "MONTANT", "FORMATEUR"]
    best_row, best_score = 0, 0
    for i, row in df_probe.iterrows():
        row_str = " ".join(str(v) for v in row.values).upper()
        score   = sum(1 for k in keywords if k in row_str)
        if score > best_score:
            best_score, best_row = score, int(i)
    return best_row


# ─────────────────────────────────────────────
# IMPUTATION
# ─────────────────────────────────────────────

def _impute_by_mode(df: pd.DataFrame, source: str = "") -> pd.DataFrame:
    """
    Impute les valeurs manquantes d'un DataFrame en se basant sur le mode de formation.

    Pour chaque colonne cible, les valeurs manquantes sont remplacées par :
      - La valeur la plus fréquente (pour les colonnes textuelles) au sein
        du même groupe "mode".
      - La médiane (pour les colonnes numériques et les dates) au sein du
        même groupe "mode".
    Un fallback global (toutes modes confondus) est appliqué si le groupe
    ne permet pas d'imputer.

    Cas particulier pour nb_jours_calc : si les dates debut/fin sont disponibles,
    la durée est d'abord recalculée directement avant d'appliquer l'imputation par médiane.

    Les valeurs résiduellement nulles après imputation sont remplacées par des
    valeurs par défaut sûres ("Non renseigné", "Inconnue").

    Paramètres
    ----------
    df : pd.DataFrame
        DataFrame des formations à compléter (modifié in place via .loc).
    source : str, optionnel
        Étiquette de contexte pour les messages de log ("fichier" ou "global").

    Retourne
    --------
    pd.DataFrame
        Le DataFrame avec les valeurs manquantes imputées.
    """
    tag = f"({source})" if source else ""

    def is_missing(series: pd.Series) -> pd.Series:
        """Détecte les valeurs manquantes y compris les chaînes vides, 'nan' et 'None'."""
        return series.isna() | series.astype(str).str.strip().isin(["", "nan", "None"])

    def fill_str(col: str) -> None:
        """Impute une colonne textuelle par la valeur la plus fréquente par groupe mode."""
        mask = is_missing(df[col])
        if mask.sum() == 0:
            return
        ref      = df[~mask].groupby("mode")[col].agg(mode_most_frequent)
        fallback = mode_most_frequent(df.loc[~mask, col])
        df.loc[mask, col] = df.loc[mask, "mode"].map(ref).fillna(fallback)
        print(f"    {mask.sum()} '{col}' imputé(s) par mode {tag}")

    def fill_num(col: str) -> None:
        """Impute une colonne numérique par la médiane par groupe mode."""
        mask = is_missing(df[col])
        if mask.sum() == 0:
            return
        ref      = df[~mask].groupby("mode")[col].median()
        fallback = df.loc[~mask, col].median()
        df.loc[mask, col] = df.loc[mask, "mode"].map(ref).fillna(fallback)
        print(f"    {mask.sum()} '{col}' imputé(s) par médiane mode {tag}")

    # Clean sentinel strings before imputation
    for sentinel in ("", "nan", "None"):
        df["formateur"] = df["formateur"].replace(sentinel, None)

    fill_str("formateur")
    fill_num("date_debut")
    fill_num("date_fin")

    # nb_jours_calc: prefer recalculation from dates
    mask_nj = is_missing(df["nb_jours_calc"])
    if mask_nj.sum() > 0:
        df.loc[mask_nj, "nb_jours_calc"] = df[mask_nj].apply(
            lambda r: max((r["date_fin"] - r["date_debut"]).days, 1)
            if pd.notna(r.get("date_debut")) and pd.notna(r.get("date_fin"))
            else None,
            axis=1,
        )
    fill_num("nb_jours_calc")
    fill_num("nb_heures")
    fill_str("etat_dossier")
    fill_str("status")

    # Fill residual nulls with safe defaults
    df["etat_dossier"] = df["etat_dossier"].fillna("Non renseigné")
    df["status"]       = df["status"].fillna("Inconnue")
    df["formateur"]    = df["formateur"].fillna("Non renseigné")

    return df


def impute_global(
    all_formations: pd.DataFrame,
    all_participants: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Imputation globale sur l'ensemble des années combinées.

    Effectue deux niveaux d'imputation après la consolidation de tous les fichiers :

    1. Imputation par mode (appel à _impute_by_mode) sur toutes les formations.

    2. Imputation cross-année du formateur : pour les formations dont le formateur
       reste manquant, recherche la valeur la plus fréquente dans le groupe
       (mode + bureau_formation) toutes années confondues, puis par mode seul
       en fallback.

    3. Imputation des affectations participants manquantes : utilise le
       bureau_formation de la formation correspondante comme fallback,
       ou "Non renseigné" si absent.

    Paramètres
    ----------
    all_formations : pd.DataFrame
        DataFrame consolidé de toutes les formations (toutes années).
    all_participants : pd.DataFrame
        DataFrame consolidé de tous les participants (toutes années).

    Retourne
    --------
    tuple (pd.DataFrame, pd.DataFrame)
        - DataFrame des formations après imputation globale.
        - DataFrame des participants après imputation des affectations.
    """
    print("\n  Imputation globale (toutes années)...")
    df = all_formations.copy()

    df = _impute_by_mode(df, source="global")

    # Cross-year formateur imputation
    mask_fmt = df["formateur"].isna() | (df["formateur"] == "Non renseigné")
    if mask_fmt.sum() > 0:
        ref_cross = (
            df[~mask_fmt]
            .groupby(["mode", "bureau_formation"])["formateur"]
            .agg(mode_most_frequent)
        )
        ref_mode = (
            df[~mask_fmt]
            .groupby("mode")["formateur"]
            .agg(mode_most_frequent)
        )

        def _impute_cross(row) -> str:
            """
            Impute le formateur d'une formation via une recherche en deux niveaux.

            Cherche d'abord dans le groupe (mode + bureau_formation), puis
            dans le groupe (mode seul) si le premier niveau ne donne rien.

            Paramètres
            ----------
            row : pd.Series
                Ligne de formation avec les champs 'mode' et 'bureau_formation'.

            Retourne
            --------
            str
                Formateur imputé ou "Non renseigné".
            """
            val = ref_cross.get((row["mode"], row["bureau_formation"]))
            if val and str(val) not in ("nan", "None", "Non renseigné"):
                return val
            return ref_mode.get(row["mode"], "Non renseigné")

        df.loc[mask_fmt, "formateur"] = df[mask_fmt].apply(_impute_cross, axis=1)
        print(f"   → {mask_fmt.sum()} formateur(s) imputé(s) cross-année")

    # Participant affectation fallback → bureau_formation
    if not all_participants.empty:
        bureau_map = df.set_index("formation_id")["bureau_formation"].to_dict()
        mask_aff   = (
            all_participants["affectation"].isna()
            | (all_participants["affectation"].astype(str).str.strip() == "")
        )
        if mask_aff.sum() > 0:
            all_participants.loc[mask_aff, "affectation"] = (
                all_participants.loc[mask_aff, "formation_id"]
                .map(bureau_map)
                .fillna("Non renseigné")
            )
            print(f"   → {mask_aff.sum()} affectation(s) participants imputée(s)")

    return df, all_participants


# ─────────────────────────────────────────────
# EXTRACT & TRANSFORM (one file)
# ─────────────────────────────────────────────

def extract_transform(filepath: Path, year: int) -> tuple[pd.DataFrame, list[dict]]:
    """
    Lit un fichier Excel brut et retourne les données nettoyées et structurées.

    Pipeline appliqué pour chaque fichier :
      1. Détection automatique de la feuille principale (detect_best_sheet).
      2. Détection automatique de la ligne d'en-tête (detect_header_row).
      3. Lecture du fichier avec l'en-tête détecté + suppression des lignes vides.
      4. Pour chaque ligne :
         - Parsing des dates (date_debut, date_fin) via clean_date.
         - Construction de l'identifiant unique formation_id = "year_numero".
         - Validation du mode via validate_mode.
         - Nettoyage du montant via clean_montant.
         - Calcul ou extraction de la durée en jours et en heures.
         - Normalisation de l'état dossier et du bureau de formation.
         - Parsing des participants via parse_participants.
      5. Déduplication sur formation_id.
      6. Imputation locale par mode via _impute_by_mode.
      7. Construction de la table participants à partir des données parsées.

    Paramètres
    ----------
    filepath : Path
        Chemin vers le fichier Excel brut à traiter.
    year : int
        Année de référence extraite du nom du fichier.

    Retourne
    --------
    tuple (pd.DataFrame, list[dict])
        - DataFrame des formations nettoyées (colonnes STANDARD_COLUMNS sans _participants).
        - Liste de dictionnaires participants {formation_id, matricule, nom, affectation, year}.
    """
    print(f"\n  {filepath.name}")

    sheet_name = detect_best_sheet(filepath)
    header_row = detect_header_row(filepath, sheet_name)

    df = pd.read_excel(filepath, sheet_name=sheet_name, header=header_row)
    df.columns = [str(c).strip() for c in df.columns]
    df.dropna(how="all", inplace=True)

    rows            : list[dict] = []
    participants_rows: list[dict] = []

    for _, series in df.iterrows():
        row = series.where(pd.notna(series), None).to_dict()

        # ── Dates ──────────────────────────────────────────────────────────
        date_debut = clean_date(get_col(row, "Date début", "Date debut", "Du", "Date de début"))
        date_fin   = clean_date(get_col(row, "Date Fin",   "Date fin",   "Au", "Date de fin"))

        # ── Formation ID ───────────────────────────────────────────────────
        numero_raw   = str(get_col(row, "Num°", "Num", "Numéro", "N° Action", "Numero") or "").strip()
        formation_id = f"{year}_{numero_raw}"

        # ── Mode ───────────────────────────────────────────────────────────
        mode = validate_mode(get_col(row, "Mode") or "", formation_id)

        # ── Montant ────────────────────────────────────────────────────────
        montant, montant_raw = clean_montant(
            get_col(row, "Montant", "Dépenses acquittées (DT) (H.TVA)")
        )

        # ── Duration (days) ────────────────────────────────────────────────
        nb_jours_src = get_col(
            row, "Nb de\n jours", "Nb jours", "Durée en Jours",
            "Durée en jours", "nb_jours_calc", "Nb de jours",
        )
        if nb_jours_src is not None:
            try:
                nb_jours_calc = int(float(str(nb_jours_src)))
            except (ValueError, TypeError):
                nb_jours_calc = None
        elif pd.notna(date_debut) and pd.notna(date_fin):
            nb_jours_calc = max((date_fin - date_debut).days, 1)
        else:
            nb_jours_calc = None

        # ── Duration (hours) ───────────────────────────────────────────────
        nb_heures_src = get_col(
            row, "Nb d'heure", "Nb d'heures", "Durée en heure",
            "Durée en Heure", "Nh", "nb_heures",
        )
        if nb_heures_src is not None:
            try:
                nb_heures = int(float(str(nb_heures_src)))
            except (ValueError, TypeError):
                nb_heures = None
        else:
            nb_heures = None

        # ── Etat dossier ───────────────────────────────────────────────────
        etat_raw     = get_col(row, "Etat de Dossier", "Etat Dossier")
        etat_dossier = normalize_etat_dossier(
            None if str(etat_raw).strip().lower() == "nan" else etat_raw
        )

        # ── Bureau formation ───────────────────────────────────────────────
        bureau_raw       = str(get_col(row, "Bureau de formation", "Structure de Formation", "Agrément") or "")
        bureau_formation = normalize_affectation(bureau_raw) or bureau_raw

        # ── Formateur ──────────────────────────────────────────────────────
        formateur = ", ".join(
            f.strip()
            for f in str(get_col(row, "Formateur") or "").replace("\r", "").split("\n")
            if f.strip()
        )

        # ── Participants ───────────────────────────────────────────────────
        participants = parse_participants(
            get_col(row, "Participant", "Participants"),
            get_col(row, "Affectation", "Affectations", "affectation"),
        )

        rows.append({
            "formation_id"    : formation_id,
            "year"            : year,
            "mode"            : mode,
            "theme"           : str(get_col(row, "Thème", "Theme", "Théme de l'action", "Thème de l'action") or "").strip(),
            "montant"         : montant,
            "montant_raw"     : montant_raw,
            "date_debut"      : date_debut,
            "date_fin"        : date_fin,
            "nb_jours_calc"   : nb_jours_calc,
            "nb_heures"       : nb_heures,
            "bureau_formation": bureau_formation,
            "formateur"       : formateur,
            "facture"         : str(get_col(row, "Facture") or ""),
            "etat_dossier"    : etat_dossier,
            "status"          : compute_status(date_debut, date_fin),
            "source_file"     : filepath.name,
            "_participants"   : participants,   # temporary — removed before export
        })

    df_clean = pd.DataFrame(rows)

    # ── Dedup ──────────────────────────────────────────────────────────────
    before = len(df_clean)
    df_clean.drop_duplicates(subset=["formation_id"], inplace=True)
    dupes = before - len(df_clean)
    if dupes:
        print(f"   {dupes} doublon(s) supprimé(s)")

    # ── Local imputation ───────────────────────────────────────────────────
    df_clean = _impute_by_mode(df_clean, source="fichier")

    # ── Build participants table ────────────────────────────────────────────
    for _, r in df_clean.iterrows():
        for p in r["_participants"]:
            aff = p.get("affectation") or r.get("bureau_formation") or "Non renseigné"
            participants_rows.append({
                "formation_id": r["formation_id"],
                "matricule"   : p["matricule"],
                "nom"         : p["nom"],
                "affectation" : aff,
                "year"        : r["year"],
            })

    df_clean.drop(columns=["_participants"], inplace=True)
    return df_clean, participants_rows


# ─────────────────────────────────────────────
# EXPORT
# ─────────────────────────────────────────────

def export_excel(df: pd.DataFrame, participants: pd.DataFrame, year: int) -> None:
    """
    Exporte les données nettoyées dans un fichier Excel clean_{year}.xlsx.

    Crée un classeur avec deux feuilles :
      - "Formations"         : une ligne par formation, colonnes STANDARD_COLUMNS.
      - "Participants_Detail" : une ligne par participant.

    Mise en forme appliquée automatiquement :
      - Les colonnes de dates (date_debut, date_fin) sont formatées en "JJ/MM/AAAA".
      - La largeur de chaque colonne est ajustée automatiquement selon le contenu
        (maximum 50 caractères).

    Seules les colonnes présentes dans le DataFrame sont exportées (évite les
    KeyError si certaines colonnes source sont absentes).

    Paramètres
    ----------
    df : pd.DataFrame
        DataFrame des formations de l'année.
    participants : pd.DataFrame
        DataFrame des participants de l'année.
    year : int
        Année utilisée pour nommer le fichier de sortie.
    """
    output    = CLEAN_FOLDER / f"clean_{year}.xlsx"
    date_cols = {"date_debut", "date_fin"}

    # Only keep columns that exist (avoids KeyError if source is missing some)
    cols_present = [c for c in STANDARD_COLUMNS if c in df.columns]

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_df, sheet_name in [
            (df[cols_present], "Formations"),
            (participants,     "Participants_Detail"),
        ]:
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]

            for col_idx, col_name in enumerate(sheet_df.columns, start=1):
                col_letter = get_column_letter(col_idx)

                # Date formatting
                if col_name in date_cols:
                    for cell in ws[col_letter][1:]:
                        cell.number_format = "DD/MM/YYYY"

                # Auto column width
                max_len = len(str(col_name))
                for cell in ws[col_letter][1:]:
                    try:
                        max_len = max(max_len, len(str(cell.value or "")))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

    print(f"   {len(df)} formations | {len(participants)} participants → {output.name}")


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main() -> None:
    """
    Point d'entrée principal du pipeline ETL.

    Orchestre les trois étapes du traitement :

    [1/3] Extract & Transform
        - Recherche tous les fichiers .xlsx dans RAW_FOLDER (hors fichiers déjà nettoyés).
        - Pour chaque fichier, extrait l'année depuis le nom du fichier (regex 4 chiffres).
        - Appelle extract_transform() sur chaque fichier et collecte les résultats.
        - Les erreurs par fichier sont capturées et affichées en fin de traitement.

    [2/3] Imputation globale
        - Consolide tous les DataFrames de formations et de participants.
        - Appelle impute_global() pour compléter les valeurs manquantes en
          exploitant l'ensemble des données multi-annuelles.

    [3/3] Export par année
        - Sépare les données consolidées par année (_year_file).
        - Appelle export_excel() pour chaque année afin de produire un fichier
          clean_{year}.xlsx dans CLEAN_FOLDER.

    Affiche un résumé final indiquant les éventuelles erreurs rencontrées.
    """
    print("=" * 55)
    print("  ETL NETTOYAGE — v2.4")
    print("=" * 55)

    files = sorted(RAW_FOLDER.glob("*.xlsx"))
    # Exclude files that are already cleaned outputs
    files = [f for f in files if not f.name.startswith("clean_")]

    if not files:
        print(f"  Aucun fichier .xlsx trouvé dans {RAW_FOLDER}")
        return

    print(f"\n[1/3] Extract & Transform ({len(files)} fichier(s))...")

    all_formations   : list[pd.DataFrame] = []
    all_participants : list[pd.DataFrame] = []
    errors           : list[tuple]        = []

    for filepath in files:
        year_match = re.search(r"(\d{4})", filepath.name)
        if not year_match:
            print(f"  ⚠  Année non détectée dans '{filepath.name}' → 2025")
        year = int(year_match.group(1)) if year_match else 2025

        try:
            df_clean, parts = extract_transform(filepath, year)
            df_clean["_year_file"] = year
            all_formations.append(df_clean)
            if parts:
                all_participants.append(pd.DataFrame(parts))
        except Exception as exc:
            print(f"  ✗ Erreur sur {filepath.name}: {exc}")
            errors.append((filepath.name, str(exc)))

    if not all_formations:
        print("  Aucune donnée extraite — arrêt.")
        return

    df_all   = pd.concat(all_formations,   ignore_index=True)
    part_all = (
        pd.concat(all_participants, ignore_index=True)
        if all_participants
        else pd.DataFrame()
    )

    print("\n[2/3] Imputation globale...")
    df_all, part_all = impute_global(df_all, part_all)

    print("\n[3/3] Export par année...")
    for year in sorted(df_all["_year_file"].unique()):
        df_year   = df_all[df_all["_year_file"] == year].drop(columns=["_year_file"])
        part_year = (
            part_all[part_all["year"] == year]
            if not part_all.empty
            else pd.DataFrame()
        )
        export_excel(df_year, part_year, int(year))

    print("\n" + "=" * 55)
    if errors:
        print(f"  Terminé avec {len(errors)} erreur(s):")
        for fname, err in errors:
            print(f"   • {fname}: {err}")
    else:
        print("  Terminé sans erreur. ✓")
    print("=" * 55)


if __name__ == "__main__":
    main()
